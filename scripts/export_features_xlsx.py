"""
Export the partner-full-features.md tables to a colour-coded multi-sheet
XLSX so the business partner can scan everything at a glance.

Layout:
  Sheet "Summary" — count by status, count by section
  One sheet per section heading from partner-full-features.md (Auth,
  Profiles, Home/Discovery, …) — each section's full feature table.
  Sheet "Out of scope" — the explicit no-go list.

Status column is conditionally coloured:
  ✅ green | ⚠️ yellow | 📋 blue | ❌ grey

Run from repo root:
    python scripts/export_features_xlsx.py
"""
from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "docs" / "business" / "partner-full-features.md"
OUT = ROOT / "docs" / "business" / "exports" / "feature-inventory.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)


# --- Style palette -----------------------------------------------------------

BRAND_RED = "E50914"
HEADER_BG = "141414"   # navbar / table header
HEADER_FG = "FFFFFF"
ZEBRA = "F7F7F7"
BORDER = Side(style="thin", color="DDDDDD")

STATUS_FILL = {
    "✅": PatternFill("solid", fgColor="D4EDDA"),  # green-ish
    "⚠️": PatternFill("solid", fgColor="FFF3CD"),  # yellow
    "📋": PatternFill("solid", fgColor="D6E5FA"),  # blue
    "❌": PatternFill("solid", fgColor="E5E5E5"),  # grey
}
STATUS_FONT = {
    "✅": Font(color="155724", bold=True),
    "⚠️": Font(color="856404", bold=True),
    "📋": Font(color="1B4F8A", bold=True),
    "❌": Font(color="333333", bold=True),
}


# --- Markdown parsing --------------------------------------------------------


def parse_sections(md_text: str) -> "OrderedDict[str, list[list[str]]]":
    """Pull `## N. Section Title` → list of rows from the markdown file.

    Each row is [#, Feature, Status, Notes] (column count varies per section
    in the source — we normalise to 4 columns here).
    """
    sections: "OrderedDict[str, list[list[str]]]" = OrderedDict()
    current_name: str | None = None
    current_rows: list[list[str]] = []
    in_table = False
    skip_separator = False
    expected_cols = 0

    for raw in md_text.splitlines():
        line = raw.rstrip()

        m = re.match(r"^##\s+(.+)$", line)
        if m:
            if current_name and current_rows:
                sections[current_name] = current_rows
            current_name = m.group(1).strip()
            current_rows = []
            in_table = False
            continue

        if not current_name:
            continue

        # Section divider — any new ##/--- pushes the current section out
        if line.startswith("---") and in_table:
            in_table = False
            continue

        # Table header
        if line.startswith("|") and not in_table:
            header_cells = [c.strip() for c in line.strip("|").split("|")]
            expected_cols = len(header_cells)
            # Remember the header so each section's sheet can name its columns
            current_rows.append(["__HEADER__"] + header_cells)
            in_table = True
            skip_separator = True
            continue

        # Table separator (---|---)
        if in_table and skip_separator and re.match(r"^\|[\s\-:|]+\|$", line):
            skip_separator = False
            continue

        # Table body row
        if in_table and line.startswith("|"):
            cells = [c.strip() for c in line.strip("|").split("|")]
            cells = (cells + [""] * expected_cols)[:expected_cols]
            current_rows.append(cells)
            continue

        # Blank line inside a table ends the table
        if in_table and not line.strip():
            in_table = False

    if current_name and current_rows:
        sections[current_name] = current_rows

    return sections


def status_symbol(cell: str) -> str | None:
    """Return the leading status emoji if the cell starts with one of our four."""
    for sym in ("✅", "⚠️", "📋", "❌"):
        if cell.startswith(sym):
            return sym
    return None


# --- Sheet builders ----------------------------------------------------------


def style_header(cell) -> None:
    cell.font = Font(bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)


def add_section_sheet(wb: Workbook, name: str, rows: list[list[str]]) -> tuple[int, dict]:
    """Returns (data_row_count, status_counts) for the summary tally."""
    # Clip sheet name to Excel's 31-char limit + strip illegal chars
    clean = re.sub(r"[\[\]:*?/\\]", "", name)[:31]
    ws = wb.create_sheet(clean)

    header = None
    data: list[list[str]] = []
    for row in rows:
        if row and row[0] == "__HEADER__":
            header = row[1:]
        else:
            data.append(row)

    if header is None:
        # No table in this section (e.g. an intro section), skip
        wb.remove(ws)
        return 0, {}

    # Find which column is the Status one (varies per section)
    status_col_idx = None
    for j, col in enumerate(header):
        if col.lower().strip() == "status":
            status_col_idx = j
            break

    # Write header
    for j, col_text in enumerate(header):
        cell = ws.cell(row=1, column=j + 1, value=col_text)
        style_header(cell)

    # Write data
    status_counts: dict[str, int] = {}
    for i, row in enumerate(data, start=2):
        for j, value in enumerate(row):
            # Strip inline markdown `code`/bold/italic for cell display
            clean_v = re.sub(r"\*\*(.+?)\*\*", r"\1", value)
            clean_v = re.sub(r"`([^`]+)`", r"\1", clean_v)
            cell = ws.cell(row=i, column=j + 1, value=clean_v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
            # Zebra striping for legibility
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)

        # Apply status colour to the Status column
        if status_col_idx is not None and status_col_idx < len(row):
            sym = status_symbol(row[status_col_idx])
            if sym:
                target = ws.cell(row=i, column=status_col_idx + 1)
                target.fill = STATUS_FILL[sym]
                target.font = STATUS_FONT[sym]
                target.alignment = Alignment(horizontal="center", vertical="center")
                status_counts[sym] = status_counts.get(sym, 0) + 1

    # Auto-ish column widths (cap at 60 to avoid Notes columns sprawling)
    col_widths = {}
    # Header lengths
    for j, col_text in enumerate(header):
        col_widths[j] = max(col_widths.get(j, 0), min(len(col_text) + 2, 30))
    # Data lengths
    for row in data:
        for j, value in enumerate(row):
            longest_line = max((len(seg) for seg in value.splitlines()), default=0)
            col_widths[j] = max(col_widths.get(j, 0), min(longest_line + 2, 60))

    # Specific column hints — first column is usually # (narrow); Notes (last) is wide
    for j, width in col_widths.items():
        col = get_column_letter(j + 1)
        if j == 0:
            ws.column_dimensions[col].width = 6
        else:
            ws.column_dimensions[col].width = max(12, width)

    # Freeze top row + add an autofilter so the partner can sort/filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Row height
    ws.row_dimensions[1].height = 28
    return len(data), status_counts


def add_summary_sheet(
    wb: Workbook, per_section: list[tuple[str, int, dict]], totals: dict
) -> None:
    """Lives at the front of the workbook with overall counts + cost notes."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="CinePile — Feature Inventory")
    title_cell.font = Font(size=20, bold=True, color=BRAND_RED)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    sub_cell = ws.cell(row=2, column=1, value="Engineering inventory for the business partner. One sheet per section.")
    sub_cell.font = Font(size=11, italic=True, color="555555")

    # Status legend (row 4-5)
    legend_row = 4
    legend = [
        ("✅ Live", "Running today, passes tests"),
        ("⚠️ Partial", "Code exists, needs polish or wiring"),
        ("📋 Planned", "Designed, not yet implemented"),
        ("❌ Out of scope", "Deliberately not building"),
    ]
    for j, (lbl, desc) in enumerate(legend):
        sym = lbl.split()[0]
        cell = ws.cell(row=legend_row, column=j + 1, value=lbl)
        cell.fill = STATUS_FILL[sym]
        cell.font = STATUS_FONT[sym]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
        ws.cell(row=legend_row + 1, column=j + 1, value=desc).font = Font(size=9, color="666666")

    # Totals (row 7-11)
    start = 7
    ws.cell(row=start, column=1, value="Overall counts").font = Font(bold=True, size=12)
    for i, sym in enumerate(("✅", "⚠️", "📋", "❌")):
        ws.cell(row=start + 1, column=i + 1, value=sym).fill = STATUS_FILL[sym]
        ws.cell(row=start + 1, column=i + 1).font = STATUS_FONT[sym]
        ws.cell(row=start + 1, column=i + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=start + 2, column=i + 1, value=totals.get(sym, 0)).font = Font(size=14, bold=True)
        ws.cell(row=start + 2, column=i + 1).alignment = Alignment(horizontal="center")
    ws.cell(row=start + 3, column=1, value=f"Total features inventoried: {sum(totals.values())}").font = Font(bold=True, color=BRAND_RED)

    # Per-section table
    h_row = start + 6
    ws.cell(row=h_row, column=1, value="Section").fill = PatternFill("solid", fgColor=HEADER_BG)
    ws.cell(row=h_row, column=2, value="✅").fill = PatternFill("solid", fgColor=HEADER_BG)
    ws.cell(row=h_row, column=3, value="⚠️").fill = PatternFill("solid", fgColor=HEADER_BG)
    ws.cell(row=h_row, column=4, value="📋").fill = PatternFill("solid", fgColor=HEADER_BG)
    ws.cell(row=h_row, column=5, value="❌").fill = PatternFill("solid", fgColor=HEADER_BG)
    ws.cell(row=h_row, column=6, value="Total").fill = PatternFill("solid", fgColor=HEADER_BG)
    for j in range(1, 7):
        cell = ws.cell(row=h_row, column=j)
        cell.font = Font(bold=True, color=HEADER_FG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
    ws.row_dimensions[h_row].height = 24

    for row_idx, (name, total, counts) in enumerate(per_section, start=h_row + 1):
        ws.cell(row=row_idx, column=1, value=name).alignment = Alignment(vertical="center")
        ws.cell(row=row_idx, column=2, value=counts.get("✅", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=counts.get("⚠️", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=counts.get("📋", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=5, value=counts.get("❌", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=total).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6).font = Font(bold=True)
        for j in range(1, 7):
            ws.cell(row=row_idx, column=j).border = Border(
                left=BORDER, right=BORDER, top=BORDER, bottom=BORDER
            )
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=j).fill = PatternFill("solid", fgColor=ZEBRA)

    ws.column_dimensions["A"].width = 40
    for c in "BCDEF":
        ws.column_dimensions[c].width = 10


# --- Driver -------------------------------------------------------------------


def main() -> None:
    md = SRC.read_text(encoding="utf-8")
    sections = parse_sections(md)

    wb = Workbook()
    # Remove the default sheet — we'll add our own ordered
    wb.remove(wb.active)

    per_section: list[tuple[str, int, dict]] = []
    totals: dict[str, int] = {}

    # Meta sections from the source markdown that aren't real feature lists
    # — exclude them from the workbook so the partner doesn't see "Summary"
    # twice or a confusing "Where to find what" file-path index.
    SKIP_SECTIONS = {"summary table", "where to find what"}

    for name, rows in sections.items():
        if name.lower().strip() in SKIP_SECTIONS:
            continue
        # Skip narrative sections that have no table
        if not any(r and r[0] == "__HEADER__" for r in rows):
            continue
        # Friendlier sheet name for the out-of-scope section (the source title
        # has a parenthetical that gets clipped by Excel's 31-char sheet-name limit)
        display_name = name
        if "Out of scope" in name:
            display_name = "Out of scope"
        count, status_counts = add_section_sheet(wb, display_name, rows)
        if count > 0:
            per_section.append((display_name, count, status_counts))
            for k, v in status_counts.items():
                totals[k] = totals.get(k, 0) + v

    add_summary_sheet(wb, per_section, totals)

    # Put Summary first
    wb.move_sheet("Summary", offset=-len(wb.sheetnames))

    wb.save(str(OUT))
    print(f"Wrote {OUT.relative_to(ROOT)}  ({OUT.stat().st_size // 1024} KB)")
    print(f"  {len(per_section)} section sheets + summary")
    # ASCII-only stdout because Windows console is cp1252 and emojis crash print()
    print(
        f"  totals: Live {totals.get(chr(0x2705), 0)}  "
        f"Partial {totals.get(chr(0x26A0) + chr(0xFE0F), 0)}  "
        f"Planned {totals.get(chr(0x1F4CB), 0)}  "
        f"Out-of-scope {totals.get(chr(0x274C), 0)}"
    )


if __name__ == "__main__":
    main()
